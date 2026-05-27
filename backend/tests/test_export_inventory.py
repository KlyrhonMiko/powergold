import asyncio
import csv
import io
from datetime import date, datetime

import pytest
from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine

from core.models.audit_log import AuditLog
from systems.admin.models.user import User
from systems.inventory.models.borrow_request import BorrowRequest
from systems.inventory.models.borrow_request_batch import BorrowRequestBatch
from systems.inventory.models.borrow_request_item import BorrowRequestItem
from systems.inventory.models.borrow_request_unit import BorrowRequestUnit
from systems.inventory.models.entrusted_item import EntrustedItem
from systems.inventory.models.inventory import InventoryItem
from systems.inventory.models.inventory_batch import InventoryBatch
from systems.inventory.models.inventory_unit import InventoryUnit
from systems.inventory.schemas.import_export_schemas import (
    CatalogExportFilters,
    CatalogExportScope,
    LedgerMovementsExportFilters,
    LedgerRequestsExportFilters,
    TimelineMode,
)
from systems.inventory.services.export_service import ExportService
from utils.time_utils import normalize_time_window


@pytest.fixture
def session() -> Session:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    with Session(engine) as db_session:
        yield db_session


@pytest.fixture
def export_service() -> ExportService:
    return ExportService()


def _consume_streaming_response(response) -> bytes:
    async def _read() -> bytes:
        chunks: list[bytes] = []
        async for chunk in response.body_iterator:
            if isinstance(chunk, bytes):
                chunks.append(chunk)
            else:
                chunks.append(chunk.encode())
        return b"".join(chunks)

    return asyncio.run(_read())


def _parse_csv_response(response) -> list[list[str]]:
    payload = _consume_streaming_response(response).decode("utf-8")
    return list(csv.reader(io.StringIO(payload)))


def _parse_workbook_response(response):
    payload = _consume_streaming_response(response)
    return load_workbook(io.BytesIO(payload))


def _seed_export_fixture_data(session: Session) -> dict[str, str]:
    released_by = User(
        first_name="Release",
        last_name="Officer",
        email="release@example.com",
        username="release-officer",
        hashed_password="hashed",
        role="staff",
        employee_id="EMP-REL-001",
    )
    returned_by = User(
        first_name="Return",
        last_name="Officer",
        email="return@example.com",
        username="return-officer",
        hashed_password="hashed",
        role="staff",
        employee_id="EMP-RET-001",
    )
    borrower_one = User(
        first_name="John",
        last_name="Doe",
        email="john.doe@example.com",
        username="john-doe",
        hashed_password="hashed",
        role="borrower",
        employee_id="EMP-2025-001",
        user_id="BOR-1001",
    )
    borrower_two = User(
        first_name="Jane",
        last_name="Smith",
        email="jane.smith@example.com",
        username="jane-smith",
        hashed_password="hashed",
        role="borrower",
        employee_id="EMP-2025-002",
        user_id="BOR-1002",
    )
    session.add_all([released_by, returned_by, borrower_one, borrower_two])
    session.flush()

    equipment = InventoryItem(
        item_id="ITEM-TRACK-001",
        name="Tracked Camera",
        category="electronics",
        classification="equipment",
        item_type="device",
        is_trackable=True,
        total_qty=2,
        available_qty=0,
    )
    materials = InventoryItem(
        item_id="ITEM-BULK-001",
        name="Cleaning Solvent",
        category="supplies",
        classification="materials",
        item_type="consumable",
        unit_of_measure="can",
        is_trackable=False,
        total_qty=75,
        available_qty=58,
    )
    session.add_all([equipment, materials])
    session.flush()

    unit_one = InventoryUnit(
        unit_id="UNIT-001",
        inventory_uuid=equipment.id,
        serial_number="SN/001",
        status="borrowed",
        condition="good",
    )
    unit_two = InventoryUnit(
        unit_id="UNIT-002",
        inventory_uuid=equipment.id,
        serial_number="SN-002",
        status="borrowed",
        condition="good",
    )
    batch_one = InventoryBatch(
        batch_id="BATCH-001",
        inventory_uuid=materials.id,
        total_qty=50,
        available_qty=40,
        status="available",
        description="Primary stock",
    )
    batch_two = InventoryBatch(
        batch_id="BATCH-002",
        inventory_uuid=materials.id,
        total_qty=25,
        available_qty=18,
        status="near_expiry",
        description="Older stock",
    )
    session.add_all([unit_one, unit_two, batch_one, batch_two])
    session.flush()

    request_one = BorrowRequest(
        request_id="REQ-001",
        transaction_ref="TXN-001",
        borrower_uuid=borrower_one.id,
        request_date=datetime(2026, 5, 1, 9, 0, 0),
        released_at=datetime(2026, 5, 1, 10, 0, 0),
        released_by=released_by.id,
        returned_at=datetime(2026, 5, 3, 15, 0, 0),
        returned_by=returned_by.id,
        received_by=returned_by.id,
        status="returned",
    )
    request_two = BorrowRequest(
        request_id="REQ-002",
        transaction_ref="TXN-002",
        borrower_uuid=borrower_two.id,
        request_date=datetime(2026, 5, 2, 11, 0, 0),
        released_at=datetime(2026, 5, 2, 12, 0, 0),
        released_by=released_by.id,
        returned_at=datetime(2026, 5, 4, 16, 0, 0),
        returned_by=returned_by.id,
        received_by=returned_by.id,
        status="returned",
    )
    session.add_all([request_one, request_two])
    session.flush()

    session.add_all(
        [
            BorrowRequestItem(borrow_uuid=request_one.id, item_uuid=equipment.id, qty_requested=1),
            BorrowRequestItem(borrow_uuid=request_one.id, item_uuid=materials.id, qty_requested=5),
            BorrowRequestItem(borrow_uuid=request_two.id, item_uuid=equipment.id, qty_requested=1),
        ]
    )
    session.flush()

    session.add_all(
        [
            BorrowRequestUnit(
                borrow_unit_id="BRU-001",
                borrow_uuid=request_one.id,
                unit_uuid=unit_one.id,
                released_by=released_by.id,
                released_at=datetime(2026, 5, 1, 10, 5, 0),
                returned_by=returned_by.id,
                returned_at=datetime(2026, 5, 3, 15, 5, 0),
                condition_on_release="good",
                condition_on_return="fair",
                return_notes="Minor scratches observed",
            ),
            BorrowRequestUnit(
                borrow_unit_id="BRU-002",
                borrow_uuid=request_two.id,
                unit_uuid=unit_two.id,
                released_by=released_by.id,
                released_at=datetime(2026, 5, 2, 12, 5, 0),
                returned_by=returned_by.id,
                returned_at=datetime(2026, 5, 4, 16, 5, 0),
                condition_on_release="excellent",
                condition_on_return="good",
            ),
        ]
    )
    session.add(
        BorrowRequestBatch(
            borrow_batch_id="BRB-001",
            borrow_uuid=request_one.id,
            batch_uuid=batch_one.id,
            qty_assigned=5,
            released_at=datetime(2026, 5, 1, 10, 10, 0),
            returned_at=datetime(2026, 5, 3, 15, 10, 0),
        )
    )
    session.commit()

    return {
        "borrower_one_user_id": borrower_one.user_id,
        "equipment_item_id": equipment.item_id,
        "serial_one": unit_one.serial_number or "",
    }


def _seed_audit_log_fixture_data(session: Session) -> None:
    actor = User(
        first_name="Audit",
        last_name="Admin",
        email="audit.admin@example.com",
        username="audit-admin",
        hashed_password="hashed",
        role="admin",
        employee_id="EMP-AUD-001",
    )
    session.add(actor)
    session.flush()

    session.add_all(
        [
            AuditLog(
                audit_id="AUD-001",
                entity_type="inventory",
                entity_id="ITEM-TRACK-001",
                action="create",
                reason_code="seed",
                actor_id=actor.id,
            ),
            AuditLog(
                audit_id="AUD-002",
                entity_type="borrow",
                entity_id="REQ-001",
                action="update",
                reason_code="release",
                actor_id=actor.id,
            ),
        ]
    )
    session.commit()


def _seed_entrusted_export_fixture_data(session: Session) -> None:
    assigned_by = User(
        first_name="Asset",
        last_name="Manager",
        email="asset.manager@example.com",
        username="asset-manager",
        hashed_password="hashed",
        role="admin",
        employee_id="EMP-ENT-001",
        user_id="ADM-ENT-001",
    )
    assigned_to = User(
        first_name="Entrusted",
        last_name="Employee",
        email="entrusted.employee@example.com",
        username="entrusted-employee",
        hashed_password="hashed",
        role="staff",
        employee_id="EMP-ENT-002",
        user_id="USR-ENT-002",
    )
    session.add_all([assigned_by, assigned_to])
    session.flush()

    item = InventoryItem(
        item_id="ITEM-ENT-001",
        name="Office Laptop",
        category="electronics",
        classification="equipment",
        item_type="device",
        is_trackable=True,
        total_qty=1,
        available_qty=0,
    )
    session.add(item)
    session.flush()

    unit = InventoryUnit(
        unit_id="UNIT-ENT-001",
        inventory_uuid=item.id,
        serial_number="LAP-001",
        status="entrusted",
        condition="excellent",
    )
    session.add(unit)
    session.flush()

    session.add(
        EntrustedItem(
            assignment_id="ENT-001",
            unit_uuid=unit.id,
            user_id=assigned_to.id,
            assigned_by=assigned_by.id,
            assigned_at=datetime(2026, 5, 6, 9, 0, 0),
            notes="Assigned for field work",
        )
    )
    session.commit()


def test_format_export_date(export_service: ExportService) -> None:
    result = export_service._format_export_date()
    assert result.count("-") == 2
    month, day, year = result.split("-")
    assert len(month) == 2
    assert len(day) == 2
    assert len(year) == 4


def test_make_export_filename(export_service: ExportService) -> None:
    filename = export_service._make_export_filename(
        "John Doe - EMP-2025-001's Borrow Request History at Monthly(May 2026)"
    )
    assert filename == "john_doe_emp_2025_001s_borrow_request_history_at_monthly_may_2026"


def test_build_borrower_label_with_user(export_service: ExportService) -> None:
    user_map = {"uuid-1": "John Doe - EMP-2025-001"}
    label = export_service._build_borrower_label(user_map, "uuid-1")
    assert label == "John Doe - EMP-2025-001"


def test_build_borrower_label_with_customer_fallback(export_service: ExportService) -> None:
    user_map: dict[str, str] = {}
    label = export_service._build_borrower_label(user_map, None, "Walk-in Customer")
    assert label == "Walk-in Customer"


def test_sanitize_sheet_title(export_service: ExportService) -> None:
    assert export_service._sanitize_sheet_title("Trackable/Equipments") == "Trackable-Equipments"
    assert export_service._sanitize_sheet_title("SN/001") == "SN-001"


def test_equipment_history_rejects_missing_item_id_v2(
    export_service: ExportService,
    session: Session,
) -> None:
    with pytest.raises(ValueError, match="item_id is required"):
        export_service._export_movements_v2(
            session=session,
            format="csv",
            movement_type=None,
            item_id=None,
            serial_number=None,
            timeline_mode=None,
            anchor_date=None,
            date_from=None,
            date_to=None,
            include_deleted=False,
            include_archived=False,
        )


def test_movement_export_filter_validates_item_id_required() -> None:
    with pytest.raises(ValueError, match="item_id is required"):
        LedgerMovementsExportFilters(format="csv", report_version="v1")


def test_borrow_history_filter_defaults_to_v2() -> None:
    filters = LedgerRequestsExportFilters(format="csv")
    assert filters.report_version == "v2"


def test_movement_export_filter_defaults_to_v2() -> None:
    filters = LedgerMovementsExportFilters(format="csv", item_id="ITEM-001")
    assert filters.report_version == "v2"


def test_borrow_history_filter_allows_specified_borrower() -> None:
    filters = LedgerRequestsExportFilters(format="csv", borrower_id="user-1")
    assert filters.borrower_id == "user-1"


def test_catalog_export_filter_defaults_to_all_scope() -> None:
    filters = CatalogExportFilters(format="csv")
    assert filters.catalog_scope == CatalogExportScope.ALL


def test_normalize_time_window_uses_weekly_start_date_range() -> None:
    window = normalize_time_window("weekly", anchor_date=date(2026, 4, 13))
    assert window.date_from == datetime(2026, 4, 13, 0, 0, 0, tzinfo=window.date_from.tzinfo)
    assert window.date_to == datetime(2026, 4, 19, 23, 59, 59, 999999, tzinfo=window.date_to.tzinfo)


def test_normalize_time_window_treats_legacy_rolling_7_day_as_weekly() -> None:
    window = normalize_time_window("rolling_7_day", anchor_date=date(2026, 4, 13))
    assert window.date_from == datetime(2026, 4, 13, 0, 0, 0, tzinfo=window.date_from.tzinfo)
    assert window.date_to == datetime(2026, 4, 19, 23, 59, 59, 999999, tzinfo=window.date_to.tzinfo)


def test_normalize_time_window_uses_selected_month_boundaries() -> None:
    window = normalize_time_window("monthly", anchor_date=date(2026, 2, 10))
    assert window.date_from == datetime(2026, 2, 1, 0, 0, 0, tzinfo=window.date_from.tzinfo)
    assert window.date_to == datetime(2026, 2, 28, 23, 59, 59, 999999, tzinfo=window.date_to.tzinfo)


def test_normalize_time_window_uses_selected_year_boundaries() -> None:
    window = normalize_time_window("yearly", anchor_date=date(2026, 7, 4))
    assert window.date_from == datetime(2026, 1, 1, 0, 0, 0, tzinfo=window.date_from.tzinfo)
    assert window.date_to == datetime(2026, 12, 31, 23, 59, 59, 999999, tzinfo=window.date_to.tzinfo)


def test_catalog_export_csv_can_be_limited_to_trackable_items(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_export_fixture_data(session)

    response = export_service.export_inventory(
        session=session,
        format="csv",
        catalog_scope=CatalogExportScope.TRACKABLE,
    )

    rows = _parse_csv_response(response)
    assert rows[0] == [
        "name",
        "category",
        "classification",
        "item_type",
        "unit_of_measure",
        "description",
        "condition",
        "quantity",
        "serial_number",
        "expiration_date",
    ]
    data_rows = rows[1:]

    assert data_rows
    assert {row[0] for row in data_rows} == {"Tracked Camera"}
    assert {row[8] for row in data_rows} == {"SN/001", "SN-002"}


def test_catalog_export_xlsx_materials_include_batch_detail_sheet(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_export_fixture_data(session)

    response = export_service.export_inventory(
        session=session,
        format="xlsx",
        catalog_scope=CatalogExportScope.NON_TRACKABLE,
    )

    workbook = _parse_workbook_response(response)
    catalog_sheet = workbook["Materials Catalog"]
    batch_sheet = workbook["Material Batches"]

    assert workbook.sheetnames == ["Materials Catalog", "Material Batches"]
    assert catalog_sheet["A1"].value == export_service._decorate_export_title("Materials Catalog")
    assert [catalog_sheet.cell(row=2, column=index).value for index in range(1, 11)] == [
        "name",
        "category",
        "classification",
        "item_type",
        "unit_of_measure",
        "description",
        "condition",
        "quantity",
        "serial_number",
        "expiration_date",
    ]
    catalog_rows = list(catalog_sheet.iter_rows(min_row=3, values_only=True))
    assert catalog_rows
    assert {row[0] for row in catalog_rows} == {"Cleaning Solvent"}
    assert {row[6] for row in catalog_rows} == {"available", "near_expiry"}
    assert {row[7] for row in catalog_rows} == {"50", "25"}

    batch_rows = list(batch_sheet.iter_rows(min_row=3, values_only=True))
    assert len(batch_rows) == 2
    batch_rows_by_id = {row[2]: row for row in batch_rows}
    assert batch_rows_by_id["BATCH-001"][:9] == ("ITEM-BULK-001", "Cleaning Solvent", "BATCH-001", "can", "available", "available", "50", "40", None)
    assert batch_rows_by_id["BATCH-001"][9]
    assert batch_rows_by_id["BATCH-001"][10] == "Primary stock"
    assert batch_rows_by_id["BATCH-002"][:9] == ("ITEM-BULK-001", "Cleaning Solvent", "BATCH-002", "can", "near_expiry", "near_expiry", "25", "18", None)
    assert batch_rows_by_id["BATCH-002"][9]
    assert batch_rows_by_id["BATCH-002"][10] == "Older stock"


def test_catalog_export_v2_summary_and_batches_reflect_material_scope(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_export_fixture_data(session)

    response = export_service.export_inventory(
        session=session,
        format="xlsx",
        report_version="v2",
        catalog_scope=CatalogExportScope.NON_TRACKABLE,
    )

    workbook = _parse_workbook_response(response)
    summary_sheet = workbook["Summary"]

    summary_values = list(summary_sheet.iter_rows(min_row=3, values_only=True))
    assert ("Report", "Materials Catalog") in summary_values
    assert ("Catalog Scope", "non_trackable") in summary_values
    assert ("Batch Rows", 2) in summary_values
    assert workbook.sheetnames == ["Summary", "Materials Catalog", "Material Batches"]


def test_catalog_export_xlsx_trackable_uses_equipment_catalog_title(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_export_fixture_data(session)

    response = export_service.export_inventory(
        session=session,
        format="xlsx",
        catalog_scope=CatalogExportScope.TRACKABLE,
    )

    workbook = _parse_workbook_response(response)
    catalog_sheet = workbook["Equipment Catalog"]

    assert workbook.sheetnames == ["Equipment Catalog"]
    assert catalog_sheet["A1"].value == export_service._decorate_export_title("Equipment Catalog")
    assert response.headers["Content-Disposition"] == "attachment; filename=equipment_catalog.xlsx"


def test_catalog_export_xlsx_all_uses_inventory_catalog_title(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_export_fixture_data(session)

    response = export_service.export_inventory(
        session=session,
        format="xlsx",
        catalog_scope=CatalogExportScope.ALL,
    )

    workbook = _parse_workbook_response(response)
    catalog_sheet = workbook["Inventory Catalog"]

    assert workbook.sheetnames == ["Inventory Catalog", "Material Batches"]
    assert catalog_sheet["A1"].value == export_service._decorate_export_title("Inventory Catalog")
    assert response.headers["Content-Disposition"] == "attachment; filename=inventory_catalog.xlsx"


def test_audit_logs_xlsx_uses_shared_single_sheet_title_and_headers(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_audit_log_fixture_data(session)

    response = export_service.export_audit_logs(
        session=session,
        format="xlsx",
    )

    workbook = _parse_workbook_response(response)
    sheet = workbook["Audit Logs Report"]

    assert workbook.sheetnames == ["Audit Logs Report"]
    assert sheet["A1"].value == export_service._decorate_export_title("Audit Logs Report")
    assert response.headers["Content-Disposition"] == "attachment; filename=audit_logs_report.xlsx"
    assert [sheet.cell(row=2, column=index).value for index in range(1, 8)] == [
        "ID",
        "Action",
        "Entity ID",
        "Entity Type",
        "Actor",
        "Timestamp",
        "Reason",
    ]


def test_entrusted_items_xlsx_uses_shared_single_sheet_title_and_headers(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_entrusted_export_fixture_data(session)

    response = export_service.export_entrusted(
        session=session,
        format="xlsx",
    )

    workbook = _parse_workbook_response(response)
    sheet = workbook["Entrusted Items Report"]

    assert workbook.sheetnames == ["Entrusted Items Report"]
    assert sheet["A1"].value == export_service._decorate_export_title("Entrusted Items Report")
    assert response.headers["Content-Disposition"] == "attachment; filename=entrusted_items_report.xlsx"
    assert [sheet.cell(row=2, column=index).value for index in range(1, 10)] == [
        "Assignment ID",
        "Item Name",
        "Serial Number",
        "Assigned To Name",
        "Assigned To ID",
        "Assigned At",
        "Returned At",
        "Status",
        "Notes",
    ]
    assert sheet.freeze_panes == "A3"


def test_audit_logs_v2_xlsx_uses_shared_summary_and_detail_titles(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_audit_log_fixture_data(session)

    response = export_service.export_audit_logs(
        session=session,
        format="xlsx",
        report_version="v2",
    )

    workbook = _parse_workbook_response(response)

    assert workbook.sheetnames == ["Summary", "Audit Logs"]
    assert workbook["Summary"]["A1"].value == export_service._decorate_export_title("Summary")
    assert workbook["Audit Logs"]["A1"].value == export_service._decorate_export_title("Audit Logs")


def test_borrow_history_csv_contains_simplified_rows(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_export_fixture_data(session)

    response = export_service._export_borrow_history_v2(
        session=session,
        format="csv",
        status=None,
        item_id=None,
        borrower_id=None,
        serial_number=None,
        timeline_mode=None,
        anchor_date=None,
        date_from=None,
        date_to=None,
        include_deleted=False,
        include_archived=False,
    )

    rows = _parse_csv_response(response)
    assert rows[0] == [
        "Request Date",
        "Request ID",
        "Borrower's Name + Employee ID",
        "What they Borrowed",
        "Serial/Batch Number",
        "Unit of Measure",
        "Quantity",
        "Returned Quantity",
        "Not Returned Quantity",
        "Return Outcome",
        "Request Status",
        "Due Date",
        "Returned On Time",
        "Released at",
        "Released by",
        "Returned at",
        "Returned by",
    ]
    assert ["05/01/2026", "REQ-001", "John Doe - EMP-2025-001", "Tracked Camera", "SN/001", "", "1", "1", "0", "Fully Returned", "returned", "", "N/A", "2026-05-01 10:05:00", "Release Officer - EMP-REL-001", "2026-05-03 15:05:00", "Return Officer - EMP-RET-001"] in rows[1:]
    assert ["05/01/2026", "REQ-001", "John Doe - EMP-2025-001", "Cleaning Solvent", "BATCH-001", "can", "5", "5", "0", "Fully Returned", "returned", "", "N/A", "2026-05-01 10:10:00", "Release Officer - EMP-REL-001", "2026-05-03 15:10:00", "Return Officer - EMP-RET-001"] in rows[1:]
    assert response.headers["Content-Disposition"] == "attachment; filename=all_borrow_request_history.csv"


def test_borrow_history_xlsx_groups_by_borrower_when_filter_is_absent(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_export_fixture_data(session)

    response = export_service._export_borrow_history_v2(
        session=session,
        format="xlsx",
        status=None,
        item_id=None,
        borrower_id=None,
        serial_number=None,
        timeline_mode=None,
        anchor_date=None,
        date_from=None,
        date_to=None,
        include_deleted=False,
        include_archived=False,
    )

    workbook = _parse_workbook_response(response)
    assert sorted(workbook.sheetnames) == ["Jane Smith - EMP-2025-002", "John Doe - EMP-2025-001"]
    borrower_sheet = workbook["John Doe - EMP-2025-001"]
    assert borrower_sheet["A1"].value == export_service._decorate_export_title("All Borrow Request History")
    assert response.headers["Content-Disposition"] == "attachment; filename=all_borrow_request_history.xlsx"
    assert [borrower_sheet.cell(row=2, column=index).value for index in range(1, 4)] == [
        "Request Date",
        "Request ID",
        "Borrower's Name + Employee ID",
    ]


def test_borrow_history_xlsx_uses_safe_sheet_names_when_borrower_is_specified(
    export_service: ExportService,
    session: Session,
) -> None:
    ids = _seed_export_fixture_data(session)

    response = export_service._export_borrow_history_v2(
        session=session,
        format="xlsx",
        status=None,
        item_id=None,
        borrower_id=ids["borrower_one_user_id"],
        serial_number=None,
        timeline_mode=None,
        anchor_date=None,
        date_from=None,
        date_to=None,
        include_deleted=False,
        include_archived=False,
    )

    workbook = _parse_workbook_response(response)
    assert workbook.sheetnames == ["Trackable-Equipments", "Untrackable-Materials"]
    assert workbook["Trackable-Equipments"]["A1"].value == export_service._decorate_export_title("John Doe - EMP-2025-001's Borrow Request History")
    assert workbook["Untrackable-Materials"]["A1"].value == export_service._decorate_export_title("John Doe - EMP-2025-001's Borrow Request History")
    assert response.headers["Content-Disposition"] == "attachment; filename=john_doe_emp_2025_001s_borrow_request_history.xlsx"


def test_borrow_history_xlsx_title_includes_timeline_context(
    export_service: ExportService,
    session: Session,
) -> None:
    _seed_export_fixture_data(session)

    response = export_service._export_borrow_history_v2(
        session=session,
        format="xlsx",
        status=None,
        item_id=None,
        borrower_id=None,
        serial_number=None,
        timeline_mode=TimelineMode.WEEKLY,
        anchor_date=date(2026, 5, 1),
        date_from=None,
        date_to=None,
        include_deleted=False,
        include_archived=False,
    )

    workbook = _parse_workbook_response(response)
    borrower_sheet = workbook["John Doe - EMP-2025-001"]
    assert borrower_sheet["A1"].value == export_service._decorate_export_title("All Borrow Request History at Weekly(May 01, 2026)")


def test_equipment_history_csv_contains_simplified_rows(
    export_service: ExportService,
    session: Session,
) -> None:
    ids = _seed_export_fixture_data(session)

    response = export_service._export_movements_v2(
        session=session,
        format="csv",
        movement_type=None,
        item_id=ids["equipment_item_id"],
        serial_number=None,
        timeline_mode=None,
        anchor_date=None,
        date_from=None,
        date_to=None,
        include_deleted=False,
        include_archived=False,
    )

    rows = _parse_csv_response(response)
    assert rows[0] == [
        "Serial Number",
        "Request ID Reference",
        "Request Date",
        "Who Borrowed",
        "Status on Release",
        "Released at",
        "Returned at",
        "Status on Return",
    ]
    assert ["SN/001", "REQ-001", "05/01/2026", "John Doe - EMP-2025-001", "good", "2026-05-01 10:05:00", "2026-05-03 15:05:00", "fair"] in rows[1:]
    assert response.headers["Content-Disposition"] == "attachment; filename=tracked_camera_equipment_history.csv"


def test_equipment_history_xlsx_groups_per_serial_with_safe_sheet_names(
    export_service: ExportService,
    session: Session,
) -> None:
    ids = _seed_export_fixture_data(session)

    response = export_service._export_movements_v2(
        session=session,
        format="xlsx",
        movement_type=None,
        item_id=ids["equipment_item_id"],
        serial_number=None,
        timeline_mode=None,
        anchor_date=None,
        date_from=None,
        date_to=None,
        include_deleted=False,
        include_archived=False,
    )

    workbook = _parse_workbook_response(response)
    assert workbook.sheetnames == ["SN-001", "SN-002"]
    serial_sheet = workbook["SN-001"]
    assert serial_sheet["A1"].value == export_service._decorate_export_title("Tracked Camera Equipment History")
    assert response.headers["Content-Disposition"] == "attachment; filename=tracked_camera_equipment_history.xlsx"
    assert [serial_sheet.cell(row=2, column=index).value for index in range(1, 4)] == [
        "Serial Number",
        "Request ID Reference",
        "Request Date",
    ]


def test_equipment_history_xlsx_title_includes_serial_and_timeline_context(
    export_service: ExportService,
    session: Session,
) -> None:
    ids = _seed_export_fixture_data(session)

    response = export_service._export_movements_v2(
        session=session,
        format="xlsx",
        movement_type=None,
        item_id=ids["equipment_item_id"],
        serial_number=ids["serial_one"],
        timeline_mode=TimelineMode.YEARLY,
        anchor_date=date(2026, 1, 1),
        date_from=None,
        date_to=None,
        include_deleted=False,
        include_archived=False,
    )

    workbook = _parse_workbook_response(response)
    serial_sheet = workbook["SN-001"]
    assert serial_sheet["A1"].value == export_service._decorate_export_title("Tracked Camera Equipment History for SN/001 at Yearly(2026)")
