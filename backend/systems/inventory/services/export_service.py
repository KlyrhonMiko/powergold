import io
import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime
from typing import List, Any, Optional
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import aliased
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlmodel import Session, select
from fastapi.responses import StreamingResponse

from core.models.audit_log import AuditLog
from systems.inventory.models.borrow_request import BorrowRequest
from systems.inventory.models.borrow_request_item import BorrowRequestItem
from systems.inventory.models.borrow_request_unit import BorrowRequestUnit
from systems.inventory.models.borrow_request_batch import BorrowRequestBatch
from systems.inventory.models.inventory import InventoryItem
from systems.inventory.models.inventory_unit import InventoryUnit
from systems.inventory.models.inventory_batch import InventoryBatch
from systems.inventory.schemas.import_export_schemas import CatalogExportScope, TimelineMode
from systems.inventory.quantity import format_quantity
from utils.time_utils import normalize_time_window

class ExportService:
    _FORMULA_PREFIXES = ("=", "+", "-", "@")
    _CONTROL_PREFIXES = ("\t", "\r", "\n")
    _INVALID_SHEET_TITLE_CHARS = set('\\/*?:[]')
    _MAX_QUERY_ROWS = 5000
    _MAX_RECEIPT_RENDER_ROWS = 1000
    _MAX_EXPORT_ITEMS = 2000
    _MAX_EXPORT_ROWS = 20000
    _RECEIPT_EXPORT_KEYS = (
        "request_id",
        "transaction_ref",
        "receipt_number",
        "status",
        "borrower_name",
        "borrower_user_id",
        "customer_name",
        "location_name",
        "released_at",
        "released_by_name",
        "expected_return_at",
        "returned_at",
        "returned_by_name",
        "is_emergency",
        "approval_channel",
        "notes",
        "items",
    )
    _RECEIPT_EXPORT_ITEM_KEYS = (
        "item_id",
        "name",
        "classification",
        "qty_released",
        "qty_returned",
        "qty_not_returned",
        "serial_numbers",
        "batch_details",
    )
    _TITLE_FILL = PatternFill(fill_type="solid", fgColor="1D4ED8")
    _HEADER_FILL = PatternFill(fill_type="solid", fgColor="DBEAFE")
    _ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")

    def _sanitize_export_cell(self, value: Any) -> Any:
        if not isinstance(value, str) or not value:
            return value

        if value.startswith(self._FORMULA_PREFIXES):
            return f"'{value}"

        # Neutralize tab/newline/carriage-return prefixed variants that spreadsheet apps may interpret.
        if value.startswith(self._CONTROL_PREFIXES):
            return f"'{value}"

        first_non_whitespace = next((char for char in value if not char.isspace()), "")
        if first_non_whitespace in self._FORMULA_PREFIXES:
            return f"'{value}"

        return value

    def _sanitize_export_rows(self, rows: List[List[Any]]) -> List[List[Any]]:
        return [
            [self._sanitize_export_cell(value) for value in row]
            for row in rows
        ]

    def _apply_visibility_filters(
        self,
        statement: Any,
        model: type[Any],
        include_deleted: bool,
        include_archived: bool,
    ) -> Any:
        if not include_deleted and hasattr(model, "is_deleted"):
            statement = statement.where(model.is_deleted.is_(False))
        if not include_archived and hasattr(model, "is_archived"):
            statement = statement.where(model.is_archived.is_(False))
        return statement

    def _has_filter_value(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return bool(value.strip())
        if isinstance(value, (list, tuple, set, dict)):
            return len(value) > 0
        return True

    def _reject_unsupported_filters(
        self,
        export_name: str,
        filters: dict[str, Any],
    ) -> None:
        unsupported = sorted(
            name for name, value in filters.items() if self._has_filter_value(value)
        )
        if unsupported:
            joined = ", ".join(unsupported)
            raise ValueError(
                f"Unsupported filter(s) for {export_name} export: {joined}."
            )

    def _resolve_time_window(
        self,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
    ) -> tuple[datetime | None, datetime | None]:
        window = normalize_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=date_from,
            date_to=date_to,
        )
        return window.date_from, window.date_to

    def _apply_datetime_window(
        self,
        statement: Any,
        column: Any,
        date_from: datetime | None,
        date_to: datetime | None,
    ) -> Any:
        if date_from:
            statement = statement.where(column >= date_from)
        if date_to:
            statement = statement.where(column <= date_to)
        return statement

    def _execute_bounded_query(
        self,
        session: Session,
        statement: Any,
        limit: int,
        entity_label: str,
    ) -> list[Any]:
        rows = session.exec(statement.limit(limit + 1)).all()
        if len(rows) > limit:
            raise ValueError(
                f"Export limit exceeded for {entity_label}. Maximum {limit} rows per export."
            )
        return rows

    def _append_inventory_row(self, rows: List[List[Any]], row: List[Any]) -> None:
        if len(rows) >= self._MAX_EXPORT_ROWS:
            raise ValueError(
                f"Export limit exceeded for inventory rows. Maximum {self._MAX_EXPORT_ROWS} rows per export."
            )
        rows.append(row)

    @staticmethod
    async def _stream_bytes(payload: bytes):
        yield payload

    def _resolve_inventory_unit_by_serial(
        self,
        session: Session,
        serial_number: str,
        include_deleted: bool,
        include_archived: bool,
    ) -> InventoryUnit | None:
        statement = select(InventoryUnit).where(InventoryUnit.serial_number == serial_number)
        statement = self._apply_visibility_filters(
            statement,
            InventoryUnit,
            include_deleted,
            include_archived,
        )
        return session.exec(statement).first()

    def _get_item_count_by_request_id(
        self,
        session: Session,
        request_ids: list[Any],
        include_deleted: bool,
        include_archived: bool,
    ) -> dict[Any, int]:
        if not request_ids:
            return {}

        item_count_statement = select(
            BorrowRequestItem.borrow_uuid,
            func.count(BorrowRequestItem.id),
        ).where(BorrowRequestItem.borrow_uuid.in_(request_ids))
        item_count_statement = self._apply_visibility_filters(
            item_count_statement,
            BorrowRequestItem,
            include_deleted,
            include_archived,
        ).group_by(BorrowRequestItem.borrow_uuid)

        return {
            borrow_uuid: count
            for borrow_uuid, count in session.exec(item_count_statement).all()
            if borrow_uuid is not None
        }

    def _get_unambiguous_borrow_request_ids_for_unit(
        self,
        session: Session,
        unit: InventoryUnit,
        include_deleted: bool,
        include_archived: bool,
    ) -> set[str]:
        request_uuid_statement = select(BorrowRequestUnit.borrow_uuid).join(
            BorrowRequest,
            BorrowRequestUnit.borrow_uuid == BorrowRequest.id,
        ).where(
            BorrowRequestUnit.unit_uuid == unit.id,
        )
        request_uuid_statement = self._apply_visibility_filters(
            request_uuid_statement,
            BorrowRequestUnit,
            include_deleted,
            include_archived,
        )
        request_uuid_statement = self._apply_visibility_filters(
            request_uuid_statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )

        request_uuids = [
            request_uuid
            for request_uuid in self._execute_bounded_query(
                session,
                request_uuid_statement,
                self._MAX_QUERY_ROWS,
                "serial-linked borrow request IDs",
            )
            if request_uuid is not None
        ]
        if not request_uuids:
            return set()

        request_unit_count_statement = (
            select(
                BorrowRequestUnit.borrow_uuid,
                func.count(func.distinct(BorrowRequestUnit.unit_uuid)),
            )
            .join(BorrowRequest, BorrowRequestUnit.borrow_uuid == BorrowRequest.id)
            .where(BorrowRequestUnit.borrow_uuid.in_(request_uuids))
            .group_by(BorrowRequestUnit.borrow_uuid)
        )
        request_unit_count_statement = self._apply_visibility_filters(
            request_unit_count_statement,
            BorrowRequestUnit,
            include_deleted,
            include_archived,
        )
        request_unit_count_statement = self._apply_visibility_filters(
            request_unit_count_statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )

        unambiguous_request_uuids = [
            borrow_uuid
            for borrow_uuid, unit_count in self._execute_bounded_query(
                session,
                request_unit_count_statement,
                self._MAX_QUERY_ROWS,
                "serial-linked borrow request unit counts",
            )
            if borrow_uuid is not None and unit_count == 1
        ]
        if not unambiguous_request_uuids:
            return set()

        request_id_statement = select(BorrowRequest.request_id).where(
            BorrowRequest.id.in_(unambiguous_request_uuids)
        )
        request_id_statement = self._apply_visibility_filters(
            request_id_statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )
        return {
            request_id
            for request_id in self._execute_bounded_query(
                session,
                request_id_statement,
                self._MAX_QUERY_ROWS,
                "serial-linked borrow request IDs",
            )
            if request_id is not None
        }

    def _format_timestamp(self, value: datetime | None) -> str:
        if value is None:
            return "N/A"
        return value.strftime("%Y-%m-%d %H:%M:%S")

    def _format_optional_timestamp(self, value: datetime | None) -> str:
        if value is None:
            return ""
        return value.strftime("%Y-%m-%d %H:%M:%S")

    def _format_bool(self, value: bool | None) -> str:
        if value is None:
            return ""
        return "Yes" if value else "No"

    def _build_user_map(self, session: Session, actor_ids: set[Any]) -> dict[Any, str]:
        from systems.admin.models.user import User

        normalized_ids = {actor_id for actor_id in actor_ids if actor_id is not None}
        if not normalized_ids:
            return {}

        users = session.exec(select(User).where(User.id.in_(normalized_ids))).all()
        return {
            user.id: f"{user.first_name} {user.last_name} - {user.employee_id or user.user_id}"
            for user in users
        }

    def _format_actor_from_map(self, user_map: dict[Any, str], actor_id: Any) -> str:
        if actor_id is None:
            return ""
        return user_map.get(actor_id, str(actor_id))

    def _days_between(self, start: datetime | None, end: datetime | None) -> str:
        if start is None or end is None:
            return ""
        return str(max((end - start).days, 0))

    def _build_borrower_label(self, user_map: dict[Any, str], actor_id: Any, customer_name: str | None = None) -> str:
        resolved = self._format_actor_from_map(user_map, actor_id)
        if resolved:
            return resolved
        customer = (customer_name or "").strip()
        if customer:
            return customer
        return "Unknown Borrower"

    def _format_export_date(self) -> str:
        now = datetime.now()
        return now.strftime("%m-%d-%Y")

    def _format_export_date_label(self) -> str:
        now = datetime.now()
        return now.strftime("%B %d, %Y")

    def _decorate_export_title(self, title: str) -> str:
        return f"{title} (Exported: {self._format_export_date_label()})"

    def _make_export_filename(self, report_title: str) -> str:
        normalized = report_title.strip().lower().replace("&", " and ")
        normalized = normalized.replace("'", "")
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
        normalized = normalized.strip("_")
        return normalized or "export"

    def _sanitize_sheet_title(self, value: str) -> str:
        sanitized = "".join("-" if char in self._INVALID_SHEET_TITLE_CHARS else char for char in value)
        sanitized = " ".join(sanitized.split()).strip(" '")
        if not sanitized:
            return "Sheet"
        return sanitized[:31]

    def _build_filter_summary_rows(
        self,
        export_name: str,
        format: str,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
        extra_filters: dict[str, Any] | None = None,
    ) -> list[list[Any]]:
        rows: list[list[Any]] = [
            ["Report", export_name],
            ["Generated At", self._format_optional_timestamp(datetime.now())],
            ["Format", format],
            ["Timeline Mode", getattr(timeline_mode, "value", timeline_mode) or ""],
            ["Anchor Date", anchor_date.isoformat() if anchor_date else ""],
            ["Date From", self._format_optional_timestamp(date_from)],
            ["Date To", self._format_optional_timestamp(date_to)],
            ["Include Deleted", self._format_bool(include_deleted)],
            ["Include Archived", self._format_bool(include_archived)],
        ]
        if extra_filters:
            rows.extend([[key, value if value is not None else ""] for key, value in extra_filters.items()])
        return rows

    def _format_export_title(self, value: str) -> str:
        return value.replace("_", " ").strip().title()

    def _format_possessive_label(self, value: str) -> str:
        stripped = value.strip()
        if not stripped:
            return stripped
        return f"{stripped}'" if stripped.lower().endswith("s") else f"{stripped}'s"

    def _format_timeline_title_suffix(
        self,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
    ) -> str:
        mode_value = getattr(timeline_mode, "value", timeline_mode)
        if mode_value == "rolling_7_day":
            mode_value = "weekly"
        if not mode_value:
            return ""

        if mode_value == "weekly":
            label = anchor_date.strftime("%B %d, %Y") if anchor_date else "Current Week"
            return f" at Weekly({label})"
        if mode_value == "monthly":
            label = anchor_date.strftime("%B %Y") if anchor_date else "Current Month"
            return f" at Monthly({label})"
        if mode_value == "yearly":
            label = anchor_date.strftime("%Y") if anchor_date else "Current Year"
            return f" at Yearly({label})"
        if mode_value == "daily":
            label = anchor_date.strftime("%B %d, %Y") if anchor_date else "Current Date"
            return f" at Daily({label})"
        return f" at {str(mode_value).title()}"

    def _build_borrow_history_title(
        self,
        borrower_label: str | None,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
    ) -> str:
        base_title = (
            "All Borrow Request History"
            if not borrower_label
            else f"{self._format_possessive_label(borrower_label)} Borrow Request History"
        )
        return f"{base_title}{self._format_timeline_title_suffix(timeline_mode, anchor_date)}"

    def _build_equipment_history_title(
        self,
        item_label: str,
        serial_number: str | None,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
    ) -> str:
        base_title = f"{item_label} Equipment History"
        if serial_number:
            base_title = f"{base_title} for {serial_number}"
        return f"{base_title}{self._format_timeline_title_suffix(timeline_mode, anchor_date)}"

    def _populate_sheet(
        self,
        ws: Any,
        title: str,
        headers: list[str],
        rows: list[list[Any]],
    ) -> None:
        title_text = self._sanitize_export_cell(self._decorate_export_title(title))
        sanitized_headers = [self._sanitize_export_cell(header) for header in headers]
        sanitized_rows = self._sanitize_export_rows(rows)
        total_columns = max(len(sanitized_headers), 1)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(bold=True, color="FFFFFF", size=14)
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = self._TITLE_FILL

        header_row = 2
        for col, header in enumerate(sanitized_headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = self._HEADER_FILL

        data_start_row = header_row + 1
        for row_offset, row_data in enumerate(sanitized_rows):
            row_number = data_start_row + row_offset
            is_alt_row = row_offset % 2 == 1
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_number, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top")
                if is_alt_row:
                    cell.fill = self._ALT_ROW_FILL

        ws.freeze_panes = f"A{data_start_row}"
        ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=header_row, column=total_columns).coordinate}"

        for column_index in range(1, total_columns + 1):
            max_length = 0
            column_letter = ws.cell(row=header_row, column=column_index).column_letter
            for cell in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=column_index,
                max_col=column_index,
            ):
                current_cell = cell[0]
                if current_cell.value is None:
                    continue
                max_length = max(max_length, len(str(current_cell.value)))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

    def _resolve_batch_condition(self, batch: InventoryBatch) -> str:
        raw_condition = getattr(batch, "condition", None)
        if raw_condition:
            return str(raw_condition)
        return batch.status or ""

    def _get_inventory_catalog_title(self, catalog_scope: CatalogExportScope) -> str:
        if catalog_scope == CatalogExportScope.TRACKABLE:
            return "Equipment Catalog"
        if catalog_scope == CatalogExportScope.NON_TRACKABLE:
            return "Materials Catalog"
        return "Inventory Catalog"

    def _collect_inventory_export_data(
        self,
        session: Session,
        catalog_scope: CatalogExportScope,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> tuple[list[InventoryItem], list[list[Any]], list[list[Any]], int, int]:
        items_statement = self._apply_visibility_filters(
            select(InventoryItem),
            InventoryItem,
            include_deleted,
            include_archived,
        )
        if catalog_scope == CatalogExportScope.TRACKABLE:
            items_statement = items_statement.where(InventoryItem.is_trackable.is_(True))
        elif catalog_scope == CatalogExportScope.NON_TRACKABLE:
            items_statement = items_statement.where(InventoryItem.is_trackable.is_(False))
        items_statement = self._apply_datetime_window(
            items_statement,
            InventoryItem.created_at,
            date_from,
            date_to,
        )
        items = self._execute_bounded_query(
            session,
            items_statement,
            self._MAX_EXPORT_ITEMS,
            "inventory items",
        )

        trackable_item_ids = [item.id for item in items if item.is_trackable]
        non_trackable_item_ids = [item.id for item in items if not item.is_trackable]

        units_by_item_id: dict[Any, list[InventoryUnit]] = defaultdict(list)
        if trackable_item_ids:
            units_statement = self._apply_visibility_filters(
                select(InventoryUnit).where(InventoryUnit.inventory_uuid.in_(trackable_item_ids)),
                InventoryUnit,
                include_deleted,
                include_archived,
            )
            units = self._execute_bounded_query(
                session,
                units_statement,
                self._MAX_EXPORT_ROWS,
                "inventory units",
            )
            for unit in units:
                units_by_item_id[unit.inventory_uuid].append(unit)

        batches_by_item_id: dict[Any, list[InventoryBatch]] = defaultdict(list)
        if non_trackable_item_ids:
            batches_statement = self._apply_visibility_filters(
                select(InventoryBatch).where(InventoryBatch.inventory_uuid.in_(non_trackable_item_ids)),
                InventoryBatch,
                include_deleted,
                include_archived,
            )
            batches = self._execute_bounded_query(
                session,
                batches_statement,
                self._MAX_EXPORT_ROWS,
                "inventory batches",
            )
            for batch in batches:
                batches_by_item_id[batch.inventory_uuid].append(batch)

        catalog_rows: list[list[Any]] = []
        material_batch_rows: list[list[Any]] = []

        for item in items:
            if item.is_trackable:
                units = units_by_item_id.get(item.id, [])
                if not units:
                    self._append_inventory_row(catalog_rows, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        "",
                        "",
                        "",
                        "0",
                        "",
                        "",
                    ])
                for unit in units:
                    self._append_inventory_row(catalog_rows, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        "",
                        unit.description or "",
                        unit.condition or "",
                        "1",
                        unit.serial_number or "",
                        unit.expiration_date.isoformat() if unit.expiration_date else "",
                    ])
            else:
                batches = batches_by_item_id.get(item.id, [])
                if not batches:
                    self._append_inventory_row(catalog_rows, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        item.unit_of_measure or "",
                        "",
                        "",
                        "0",
                        "",
                        "",
                    ])
                for batch in batches:
                    batch_condition = self._resolve_batch_condition(batch)
                    self._append_inventory_row(catalog_rows, [
                        item.name,
                        item.category or "",
                        item.classification or "",
                        item.item_type or "",
                        item.unit_of_measure or "",
                        batch.description or "",
                        batch_condition,
                        format_quantity(batch.total_qty),
                        "",
                        batch.expiration_date.isoformat() if batch.expiration_date else "",
                    ])
                    material_batch_rows.append([
                        item.item_id,
                        item.name,
                        batch.batch_id,
                        item.unit_of_measure or "",
                        batch_condition,
                        batch.status or "",
                        format_quantity(batch.total_qty),
                        format_quantity(batch.available_qty),
                        batch.expiration_date.isoformat() if batch.expiration_date else "",
                        self._format_optional_timestamp(batch.received_at),
                        batch.description or "",
                    ])

        unit_row_count = sum(len(units) for units in units_by_item_id.values())
        batch_row_count = sum(len(batches) for batches in batches_by_item_id.values())
        return items, catalog_rows, material_batch_rows, unit_row_count, batch_row_count

    def _create_multi_sheet_response(
        self,
        sheets: list[tuple[str, list[str], list[list[Any]]] | tuple[str, list[str], list[list[Any]], str]],
        filename_title: str,
    ) -> StreamingResponse:
        filename = self._make_export_filename(filename_title)
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_spec in sheets:
            if len(sheet_spec) == 4:
                sheet_name, headers, rows, display_title = sheet_spec
            else:
                sheet_name, headers, rows = sheet_spec
                display_title = sheet_name
            ws = wb.create_sheet(title=self._sanitize_sheet_title(sheet_name))
            self._populate_sheet(ws, display_title, headers, rows)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            self._stream_bytes(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    def _build_receipt_export_payload(self, receipt_payload: dict[str, Any]) -> dict[str, Any]:
        export_payload: dict[str, Any] = {
            key: receipt_payload.get(key)
            for key in self._RECEIPT_EXPORT_KEYS
            if key in receipt_payload and key != "items"
        }

        raw_items = receipt_payload.get("items")
        export_payload["items"] = [
            {
                item_key: item_payload.get(item_key)
                for item_key in self._RECEIPT_EXPORT_ITEM_KEYS
                if item_key in item_payload
            }
            for item_payload in raw_items
            if isinstance(item_payload, dict)
        ] if isinstance(raw_items, list) else []

        return export_payload

    def _get_receipt_rendered_payload(self, receipt_service: Any, session: Session, request_id: str) -> str:
        try:
            receipt = receipt_service.generate_release_receipt(session, request_id)
        except ValueError:
            return ""

        if hasattr(receipt, "model_dump"):
            receipt_payload = receipt.model_dump(mode="json")
        else:
            receipt_payload = receipt

        if isinstance(receipt_payload, dict):
            return json.dumps(self._build_receipt_export_payload(receipt_payload))
        return json.dumps(receipt_payload)

    def export_audit_logs(
        self, 
        session: Session, 
        format: str, 
        report_version: str = "v1",
        from_date: Optional[datetime] = None, 
        to_date: Optional[datetime] = None,
        timeline_mode: TimelineMode | None = None,
        anchor_date: date | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        include_deleted: bool = False,
        include_archived: bool = False,
    ) -> StreamingResponse:
        effective_from_date = date_from or from_date
        effective_to_date = date_to or to_date
        normalized_from_date, normalized_to_date = self._resolve_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=effective_from_date,
            date_to=effective_to_date,
        )

        if report_version == "v2":
            return self._export_audit_logs_v2(
                session=session,
                format=format,
                timeline_mode=timeline_mode,
                anchor_date=anchor_date,
                date_from=normalized_from_date,
                date_to=normalized_to_date,
                include_deleted=include_deleted,
                include_archived=include_archived,
            )

        statement = select(AuditLog)
        statement = self._apply_visibility_filters(
            statement,
            AuditLog,
            include_deleted,
            include_archived,
        )
        statement = self._apply_datetime_window(
            statement,
            AuditLog.created_at,
            normalized_from_date,
            normalized_to_date,
        )
            
        logs = self._execute_bounded_query(
            session,
            statement,
            self._MAX_QUERY_ROWS,
            "audit logs",
        )
        
        headers = ["ID", "Action", "Entity ID", "Entity Type", "Actor", "Timestamp", "Reason"]
        data = [
            [
                str(log.id),
                log.action,
                log.entity_id,
                log.entity_type,
                str(log.actor_id),
                log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                log.reason_code or ""
            ]
            for log in logs
        ]

        return self._create_response(headers, data, format, "Audit Logs Report")

    def export_inventory(
        self, 
        session: Session, 
        format: str,
        report_version: str = "v1",
        catalog_scope: CatalogExportScope = CatalogExportScope.ALL,
        timeline_mode: TimelineMode | None = None,
        anchor_date: date | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        include_deleted: bool = False,
        include_archived: bool = False,
    ) -> StreamingResponse:
        normalized_from_date, normalized_to_date = self._resolve_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=date_from,
            date_to=date_to,
        )

        if report_version == "v2":
            return self._export_inventory_v2(
                session=session,
                format=format,
                catalog_scope=catalog_scope,
                timeline_mode=timeline_mode,
                anchor_date=anchor_date,
                date_from=normalized_from_date,
                date_to=normalized_to_date,
                include_deleted=include_deleted,
                include_archived=include_archived,
            )

        headers = ["name", "category", "classification", "item_type", "unit_of_measure", "description", "condition", "quantity", "serial_number", "expiration_date"]
        batch_headers = [
            "item_id",
            "name",
            "batch_id",
            "unit_of_measure",
            "condition",
            "status",
            "total_quantity",
            "available_quantity",
            "expiration_date",
            "received_at",
            "description",
        ]
        items, catalog_rows, material_batch_rows, _, _ = self._collect_inventory_export_data(
            session=session,
            catalog_scope=catalog_scope,
            date_from=normalized_from_date,
            date_to=normalized_to_date,
            include_deleted=include_deleted,
            include_archived=include_archived,
        )
        catalog_title = self._get_inventory_catalog_title(catalog_scope)

        if format == "csv":
            return self._create_response(headers, catalog_rows, format, catalog_title)

        sheets: list[tuple[str, list[str], list[list[Any]]]] = [
            (catalog_title, headers, catalog_rows),
        ]
        if any(not item.is_trackable for item in items):
            sheets.append(("Material Batches", batch_headers, material_batch_rows))
        return self._create_multi_sheet_response(sheets, catalog_title)

    def export_borrow_history(
        self, 
        session: Session, 
        format: str, 
        report_version: str = "v2",
        status: Optional[str] = None,
        item_id: Optional[str] = None,
        borrower_id: Optional[str] = None,
        serial_number: Optional[str] = None,
        include_receipt_rendered: bool = False,
        timeline_mode: TimelineMode | None = None,
        anchor_date: date | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        include_deleted: bool = False,
        include_archived: bool = False,
    ) -> StreamingResponse:
        from systems.admin.models.user import User

        normalized_from_date, normalized_to_date = self._resolve_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=date_from,
            date_to=date_to,
        )

        if report_version != "v1":
            return self._export_borrow_history_v2(
                session=session,
                format=format,
                status=status,
                item_id=item_id,
                borrower_id=borrower_id,
                serial_number=serial_number,
                timeline_mode=timeline_mode,
                anchor_date=anchor_date,
                date_from=normalized_from_date,
                date_to=normalized_to_date,
                include_deleted=include_deleted,
                include_archived=include_archived,
            )

        selected_borrower_label: str | None = None
        if borrower_id:
            selected_borrower = session.exec(
                select(User).where(User.user_id == borrower_id)
            ).first()
            if selected_borrower is not None:
                selected_borrower_label = (
                    f"{selected_borrower.first_name} {selected_borrower.last_name} - "
                    f"{selected_borrower.employee_id or selected_borrower.user_id}"
                )
            else:
                selected_borrower_label = borrower_id

        report_title = self._build_borrow_history_title(
            borrower_label=selected_borrower_label,
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
        )

        headers = [
            "Request ID",
            "Borrower Name + Employee ID",
            "Item Name",
            "Serial Number",
            "Who approved the request",
            "Who assigned the unit",
            "Unit CONDITION on release",
            "Who released",
            "When released",
            "When returned",
            "Who received the return",
            "Unit CONDITION on return",
        ]
        receipt_service: Any | None = None
        receipt_payload_cache: dict[str, str] = {}
        if include_receipt_rendered:
            from systems.inventory.services.borrow_request_service import BorrowService

            receipt_service = BorrowService()

        query_row_limit = (
            self._MAX_RECEIPT_RENDER_ROWS
            if include_receipt_rendered
            else self._MAX_QUERY_ROWS
        )

        def get_cached_receipt_payload(request_id_value: str) -> str:
            if receipt_service is None:
                return ""
            if request_id_value not in receipt_payload_cache:
                receipt_payload_cache[request_id_value] = self._get_receipt_rendered_payload(
                    receipt_service,
                    session,
                    request_id_value,
                )
            return receipt_payload_cache[request_id_value]
        if receipt_service:
            headers.append("Receipt Rendered")

        normalized_serial_number = serial_number.strip() if serial_number else None

        borrower_user = aliased(User)
        approved_user = aliased(User)
        assigned_user = aliased(User)
        assignment_released_user = aliased(User)
        request_released_user = aliased(User)
        received_return_user = aliased(User)

        statement = (
            select(
                BorrowRequestUnit,
                BorrowRequest,
                InventoryUnit,
                InventoryItem,
                borrower_user,
                approved_user,
                assigned_user,
                assignment_released_user,
                request_released_user,
                received_return_user,
            )
            .join(BorrowRequest, BorrowRequestUnit.borrow_uuid == BorrowRequest.id)
            .outerjoin(InventoryUnit, BorrowRequestUnit.unit_uuid == InventoryUnit.id)
            .outerjoin(InventoryItem, InventoryUnit.inventory_uuid == InventoryItem.id)
            .outerjoin(borrower_user, BorrowRequest.borrower_uuid == borrower_user.id)
            .outerjoin(approved_user, BorrowRequest.approved_by == approved_user.id)
            .outerjoin(assigned_user, BorrowRequestUnit.assigned_by == assigned_user.id)
            .outerjoin(
                assignment_released_user,
                BorrowRequestUnit.released_by == assignment_released_user.id,
            )
            .outerjoin(
                request_released_user,
                BorrowRequest.released_by == request_released_user.id,
            )
            .outerjoin(received_return_user, BorrowRequest.received_by == received_return_user.id)
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequestUnit,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            InventoryUnit,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            InventoryItem,
            include_deleted,
            include_archived,
        )
        statement = self._apply_datetime_window(
            statement,
            BorrowRequest.request_date,
            normalized_from_date,
            normalized_to_date,
        )

        if status and status != "all":
            statement = statement.where(BorrowRequest.status == status)

        if item_id:
            statement = statement.where(InventoryItem.item_id == item_id)

        if borrower_id:
            statement = statement.where(borrower_user.user_id == borrower_id)

        if normalized_serial_number:
            statement = statement.where(InventoryUnit.serial_number == normalized_serial_number)

        results = self._execute_bounded_query(
            session,
            statement,
            query_row_limit,
            "borrow history",
        )

        def _format_actor(user_obj: User | None) -> str:
            if user_obj is None:
                return "N/A"
            return f"{user_obj.first_name} {user_obj.last_name} ({user_obj.employee_id or user_obj.user_id})"

        data = [
            [
                req.request_id,
                _format_actor(borrower) if borrower else (req.customer_name or "N/A"),
                inventory_item.name if inventory_item else "Unknown Item",
                unit.serial_number if unit and unit.serial_number else "",
                _format_actor(approved),
                _format_actor(assigned),
                assignment.condition_on_release
                or (unit.condition if unit and unit.condition else "N/A"),
                _format_actor(assignment_released or request_released),
                self._format_timestamp(assignment.released_at or req.released_at),
                self._format_timestamp(assignment.returned_at or req.returned_at),
                _format_actor(received_return),
                assignment.condition_on_return or "N/A",
                *(
                    [get_cached_receipt_payload(req.request_id)]
                    if receipt_service
                    else []
                ),
            ]
            for (
                assignment,
                req,
                unit,
                inventory_item,
                borrower,
                approved,
                assigned,
                assignment_released,
                request_released,
                received_return,
            ) in results
        ]

        return self._create_response(headers, data, format, report_title)

    def export_movements(
        self, 
        session: Session, 
        format: str, 
        report_version: str = "v2",
        movement_type: Optional[str] = None,
        item_id: Optional[str] = None,
        serial_number: Optional[str] = None,
        timeline_mode: TimelineMode | None = None,
        anchor_date: date | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        include_deleted: bool = False,
        include_archived: bool = False,
    ) -> StreamingResponse:
        normalized_from_date, normalized_to_date = self._resolve_time_window(
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
            date_from=date_from,
            date_to=date_to,
        )

        if report_version != "v1":
            return self._export_movements_v2(
                session=session,
                format=format,
                movement_type=movement_type,
                item_id=item_id,
                serial_number=serial_number,
                timeline_mode=timeline_mode,
                anchor_date=anchor_date,
                date_from=normalized_from_date,
                date_to=normalized_to_date,
                include_deleted=include_deleted,
                include_archived=include_archived,
            )

        selected_item = None
        selected_item_label = item_id or "Equipment"
        if item_id:
            selected_item = session.exec(
                self._apply_visibility_filters(
                    select(InventoryItem).where(InventoryItem.item_id == item_id),
                    InventoryItem,
                    include_deleted,
                    include_archived,
                )
            ).first()
            if selected_item and selected_item.name:
                selected_item_label = selected_item.name

        report_title = self._build_equipment_history_title(
            item_label=selected_item_label,
            serial_number=serial_number.strip() if serial_number else None,
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
        )

        headers = [
            "Serial Number",
            "Item Name",
            "Who Borrowed",
            "Who Released",
            "When It was Borrowed",
            "Condition on Release",
            "When It was Returned",
            "Who received the return",
            "Condition on Return",
            "Return Notes",
            "Request ID (As a Reference)",
        ]
        from systems.admin.models.user import User

        normalized_serial_number = serial_number.strip() if serial_number else None
        normalized_movement_type = (movement_type or "all").strip().lower()
        if normalized_movement_type not in {"all", "out", "in"}:
            raise ValueError("movement_type must be one of: all, out, in")

        borrower_user = aliased(User)
        assignment_released_user = aliased(User)
        request_released_user = aliased(User)
        received_return_user = aliased(User)
        borrowed_timestamp = func.coalesce(
            BorrowRequestUnit.released_at,
            BorrowRequest.released_at,
        )
        borrowed_window_timestamp = func.coalesce(
            BorrowRequestUnit.released_at,
            BorrowRequest.released_at,
            BorrowRequest.request_date,
        )
        returned_timestamp = func.coalesce(
            BorrowRequestUnit.returned_at,
            BorrowRequest.returned_at,
        )

        statement = (
            select(
                BorrowRequestUnit,
                BorrowRequest,
                InventoryUnit,
                InventoryItem,
                borrower_user,
                assignment_released_user,
                request_released_user,
                received_return_user,
            )
            .join(BorrowRequest, BorrowRequestUnit.borrow_uuid == BorrowRequest.id)
            .outerjoin(InventoryUnit, BorrowRequestUnit.unit_uuid == InventoryUnit.id)
            .outerjoin(InventoryItem, InventoryUnit.inventory_uuid == InventoryItem.id)
            .outerjoin(borrower_user, BorrowRequest.borrower_uuid == borrower_user.id)
            .outerjoin(
                assignment_released_user,
                BorrowRequestUnit.released_by == assignment_released_user.id,
            )
            .outerjoin(
                request_released_user,
                BorrowRequest.released_by == request_released_user.id,
            )
            .outerjoin(received_return_user, BorrowRequest.received_by == received_return_user.id)
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequestUnit,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            BorrowRequest,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            InventoryUnit,
            include_deleted,
            include_archived,
        )
        statement = self._apply_visibility_filters(
            statement,
            InventoryItem,
            include_deleted,
            include_archived,
        )

        if normalized_movement_type == "out":
            statement = statement.where(borrowed_timestamp.is_not(None))
            statement = self._apply_datetime_window(
                statement,
                borrowed_timestamp,
                normalized_from_date,
                normalized_to_date,
            )
        elif normalized_movement_type == "in":
            statement = statement.where(returned_timestamp.is_not(None))
            statement = self._apply_datetime_window(
                statement,
                returned_timestamp,
                normalized_from_date,
                normalized_to_date,
            )
        else:
            if normalized_from_date or normalized_to_date:
                borrowed_window_predicates = []
                returned_window_predicates = []

                if normalized_from_date:
                    borrowed_window_predicates.append(
                        borrowed_window_timestamp >= normalized_from_date
                    )
                    returned_window_predicates.append(
                        returned_timestamp >= normalized_from_date
                    )
                if normalized_to_date:
                    borrowed_window_predicates.append(
                        borrowed_window_timestamp <= normalized_to_date
                    )
                    returned_window_predicates.append(
                        returned_timestamp <= normalized_to_date
                    )

                statement = statement.where(
                    or_(
                        and_(*borrowed_window_predicates),
                        and_(*returned_window_predicates),
                    )
                )

        if item_id:
            statement = statement.where(InventoryItem.item_id == item_id)

        if normalized_serial_number:
            statement = statement.where(InventoryUnit.serial_number == normalized_serial_number)

        results = self._execute_bounded_query(
            session,
            statement,
            self._MAX_QUERY_ROWS,
            "equipment history",
        )

        def _format_actor(user_obj: User | None) -> str:
            if user_obj is None:
                return "N/A"
            return f"{user_obj.first_name} {user_obj.last_name} ({user_obj.employee_id or user_obj.user_id})"

        data = [
            [
                unit.serial_number if unit and unit.serial_number else "",
                item.name if item else "Unknown Item",
                _format_actor(borrower) if borrower else (req.customer_name or "N/A"),
                _format_actor(assignment_released or request_released),
                self._format_timestamp(assignment.released_at or req.released_at or req.request_date),
                assignment.condition_on_release
                or (unit.condition if unit and unit.condition else "N/A"),
                self._format_timestamp(assignment.returned_at or req.returned_at),
                _format_actor(received_return),
                assignment.condition_on_return or "N/A",
                assignment.return_notes or "",
                req.request_id,
            ]
            for (
                assignment,
                req,
                unit,
                item,
                borrower,
                assignment_released,
                request_released,
                received_return,
            ) in results
        ]

        return self._create_response(headers, data, format, report_title)

    def _export_audit_logs_v2(
        self,
        session: Session,
        format: str,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> StreamingResponse:
        statement = select(AuditLog)
        statement = self._apply_visibility_filters(
            statement,
            AuditLog,
            include_deleted,
            include_archived,
        )
        statement = self._apply_datetime_window(statement, AuditLog.created_at, date_from, date_to)
        logs = self._execute_bounded_query(session, statement, self._MAX_QUERY_ROWS, "audit logs")

        user_map = self._build_user_map(session, {log.actor_id for log in logs})
        headers = [
            "ID",
            "Audit ID",
            "Action",
            "Entity ID",
            "Entity Type",
            "Actor UUID",
            "Actor",
            "Timestamp",
            "Reason",
        ]
        rows = [
            [
                str(log.id),
                log.audit_id,
                log.action,
                log.entity_id,
                log.entity_type,
                str(log.actor_id) if log.actor_id else "",
                self._format_actor_from_map(user_map, log.actor_id),
                self._format_optional_timestamp(log.created_at),
                log.reason_code or "",
            ]
            for log in logs
        ]

        report_title = "Audit Logs Report"

        if format == "csv":
            return self._create_response(headers, rows, format, report_title)

        summary_rows = self._build_filter_summary_rows(
            "Audit Logs",
            format,
            timeline_mode,
            anchor_date,
            date_from,
            date_to,
            include_deleted,
            include_archived,
            {"Total Logs": len(logs)},
        )
        for label, values in (
            ("Action", [log.action for log in logs]),
            ("Entity Type", [log.entity_type for log in logs]),
            ("Reason", [log.reason_code or "None" for log in logs]),
        ):
            counts: dict[str, int] = defaultdict(int)
            for value in values:
                counts[value] += 1
            for value, count in sorted(counts.items()):
                summary_rows.append([f"{label}: {value}", count])

        return self._create_multi_sheet_response(
            [
                ("Summary", ["Metric", "Value"], summary_rows),
                ("Audit Logs", headers, rows),
            ],
            report_title,
        )

    def _export_inventory_v2(
        self,
        session: Session,
        format: str,
        catalog_scope: CatalogExportScope,
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> StreamingResponse:
        headers = [
            "name",
            "category",
            "classification",
            "item_type",
            "unit_of_measure",
            "description",
            "condition",
            "quantity",
            "serial_number",
            "expiration_date",
        ]
        batch_headers = [
            "item_id",
            "name",
            "batch_id",
            "unit_of_measure",
            "condition",
            "status",
            "total_quantity",
            "available_quantity",
            "expiration_date",
            "received_at",
            "description",
        ]
        items, rows, material_batch_rows, unit_row_count, batch_row_count = self._collect_inventory_export_data(
            session=session,
            catalog_scope=catalog_scope,
            date_from=date_from,
            date_to=date_to,
            include_deleted=include_deleted,
            include_archived=include_archived,
        )
        catalog_title = self._get_inventory_catalog_title(catalog_scope)

        if format == "csv":
            return self._create_response(headers, rows, format, catalog_title)

        summary_rows = self._build_filter_summary_rows(
            catalog_title,
            format,
            timeline_mode,
            anchor_date,
            date_from,
            date_to,
            include_deleted,
            include_archived,
            {
                "Catalog Scope": catalog_scope.value,
                "Total Items": len(items),
                "Trackable Items": sum(1 for item in items if item.is_trackable),
                "Non-Trackable Items": sum(1 for item in items if not item.is_trackable),
                "Unit Rows": unit_row_count,
                "Batch Rows": batch_row_count,
            },
        )
        for label, values in (
            ("Category", [item.category or "None" for item in items]),
            ("Classification", [item.classification or "None" for item in items]),
            ("Item Type", [item.item_type or "None" for item in items]),
        ):
            counts: dict[str, int] = defaultdict(int)
            for value in values:
                counts[value] += 1
            for value, count in sorted(counts.items()):
                summary_rows.append([f"{label}: {value}", count])

        sheets: list[tuple[str, list[str], list[list[Any]]]] = [
            ("Summary", ["Metric", "Value"], summary_rows),
            (catalog_title, headers, rows),
        ]
        if any(not item.is_trackable for item in items):
            sheets.append(("Material Batches", batch_headers, material_batch_rows))

        return self._create_multi_sheet_response(sheets, catalog_title)

    def _export_borrow_history_v2(
        self,
        session: Session,
        format: str,
        status: Optional[str],
        item_id: Optional[str],
        borrower_id: Optional[str],
        serial_number: Optional[str],
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> StreamingResponse:
        from systems.admin.models.user import User

        borrower_user = aliased(User)
        requested_item = aliased(InventoryItem)
        serial_unit = aliased(InventoryUnit)
        normalized_serial_number = serial_number.strip() if serial_number else None
        selected_borrower_label: str | None = None

        if borrower_id:
            selected_borrower = session.exec(
                select(User).where(User.user_id == borrower_id)
            ).first()
            if selected_borrower is not None:
                selected_borrower_label = (
                    f"{selected_borrower.first_name} {selected_borrower.last_name} - "
                    f"{selected_borrower.employee_id or selected_borrower.user_id}"
                )
            else:
                selected_borrower_label = borrower_id

        statement = select(BorrowRequest.id, BorrowRequest.request_date).outerjoin(
            borrower_user,
            BorrowRequest.borrower_uuid == borrower_user.id,
        )
        statement = self._apply_visibility_filters(statement, BorrowRequest, include_deleted, include_archived)
        statement = self._apply_datetime_window(statement, BorrowRequest.request_date, date_from, date_to)

        if status and status != "all":
            statement = statement.where(BorrowRequest.status == status)
        if borrower_id:
            statement = statement.where(borrower_user.user_id == borrower_id)
        if item_id:
            statement = statement.join(
                BorrowRequestItem, BorrowRequestItem.borrow_uuid == BorrowRequest.id
            ).join(
                requested_item, BorrowRequestItem.item_uuid == requested_item.id
            )
            statement = self._apply_visibility_filters(statement, BorrowRequestItem, include_deleted, include_archived)
            statement = self._apply_visibility_filters(statement, requested_item, include_deleted, include_archived)
            statement = statement.where(requested_item.item_id == item_id)
        if normalized_serial_number:
            statement = statement.join(
                BorrowRequestUnit, BorrowRequestUnit.borrow_uuid == BorrowRequest.id
            ).join(
                serial_unit, BorrowRequestUnit.unit_uuid == serial_unit.id
            )
            statement = self._apply_visibility_filters(statement, BorrowRequestUnit, include_deleted, include_archived)
            statement = self._apply_visibility_filters(statement, serial_unit, include_deleted, include_archived)
            statement = statement.where(serial_unit.serial_number == normalized_serial_number)

        request_keys = self._execute_bounded_query(
            session,
            statement.distinct().order_by(BorrowRequest.request_date.desc()),
            self._MAX_QUERY_ROWS,
            "borrow history requests",
        )
        request_ids = [request_id for request_id, _ in request_keys if request_id is not None]
        requests_ordered: list[BorrowRequest] = []
        if request_ids:
            request_statement = self._apply_visibility_filters(
                select(BorrowRequest).where(BorrowRequest.id.in_(request_ids)),
                BorrowRequest,
                include_deleted,
                include_archived,
            )
            request_map = {
                request.id: request
                for request in self._execute_bounded_query(
                    session,
                    request_statement,
                    self._MAX_QUERY_ROWS,
                    "borrow history request records",
                )
            }
            requests_ordered = [request_map[rid] for rid in request_ids if rid in request_map]

        unit_rows_by_request_id: dict[Any, list[tuple[BorrowRequestUnit, InventoryUnit | None, InventoryItem | None]]] = defaultdict(list)
        batch_rows_by_request_id: dict[Any, list[tuple[BorrowRequestBatch, InventoryBatch | None, InventoryItem | None]]] = defaultdict(list)
        request_items_by_request_id: dict[Any, list[BorrowRequestItem]] = defaultdict(list)
        item_by_id: dict[Any, InventoryItem] = {}

        if request_ids:
            item_statement = self._apply_visibility_filters(
                select(BorrowRequestItem).where(BorrowRequestItem.borrow_uuid.in_(request_ids)),
                BorrowRequestItem, include_deleted, include_archived,
            )
            request_items = self._execute_bounded_query(
                session, item_statement.order_by(BorrowRequestItem.created_at.asc()),
                self._MAX_EXPORT_ROWS, "borrow request items",
            )
            item_uuids = {row.item_uuid for row in request_items if row.item_uuid is not None}
            if item_uuids:
                inv_statement = self._apply_visibility_filters(
                    select(InventoryItem).where(InventoryItem.id.in_(item_uuids)),
                    InventoryItem, include_deleted, include_archived,
                )
                item_by_id = {
                    item.id: item
                    for item in self._execute_bounded_query(session, inv_statement, self._MAX_EXPORT_ITEMS, "borrow request item details")
                }
            for request_item in request_items:
                request_items_by_request_id[request_item.borrow_uuid].append(request_item)

            unit_statement = (
                select(BorrowRequestUnit, InventoryUnit, InventoryItem)
                .outerjoin(InventoryUnit, BorrowRequestUnit.unit_uuid == InventoryUnit.id)
                .outerjoin(InventoryItem, InventoryUnit.inventory_uuid == InventoryItem.id)
                .where(BorrowRequestUnit.borrow_uuid.in_(request_ids))
            )
            unit_statement = self._apply_visibility_filters(unit_statement, BorrowRequestUnit, include_deleted, include_archived)
            unit_statement = self._apply_visibility_filters(unit_statement, InventoryUnit, include_deleted, include_archived)
            unit_statement = self._apply_visibility_filters(unit_statement, InventoryItem, include_deleted, include_archived)
            unit_rows = self._execute_bounded_query(session, unit_statement, self._MAX_EXPORT_ROWS, "borrow request unit assignments")
            for assignment, unit, inventory_item in unit_rows:
                unit_rows_by_request_id[assignment.borrow_uuid].append((assignment, unit, inventory_item))

            batch_statement = (
                select(BorrowRequestBatch, InventoryBatch, InventoryItem)
                .outerjoin(InventoryBatch, BorrowRequestBatch.batch_uuid == InventoryBatch.id)
                .outerjoin(InventoryItem, InventoryBatch.inventory_uuid == InventoryItem.id)
                .where(BorrowRequestBatch.borrow_uuid.in_(request_ids))
            )
            batch_statement = self._apply_visibility_filters(batch_statement, BorrowRequestBatch, include_deleted, include_archived)
            batch_statement = self._apply_visibility_filters(batch_statement, InventoryBatch, include_deleted, include_archived)
            batch_statement = self._apply_visibility_filters(batch_statement, InventoryItem, include_deleted, include_archived)
            batch_rows = self._execute_bounded_query(session, batch_statement, self._MAX_EXPORT_ROWS, "borrow request batch assignments")
            for assignment, batch, inventory_item in batch_rows:
                batch_rows_by_request_id[assignment.borrow_uuid].append((assignment, batch, inventory_item))

        actor_ids: set[Any] = set()
        for request in requests_ordered:
            actor_ids.update({request.borrower_uuid, request.approved_by, request.released_by, request.returned_by, request.received_by})
        for rows in unit_rows_by_request_id.values():
            for assignment, _, _ in rows:
                actor_ids.update({assignment.assigned_by, assignment.released_by, assignment.returned_by})
        user_map = self._build_user_map(session, actor_ids)

        headers = [
            "Request Date",
            "Request ID",
            "Borrower's Name + Employee ID",
            "What they Borrowed",
            "Serial/Batch Number",
            "Unit of Measure",
            "Quantity",
            "Returned Quantity",
            "Not Returned Quantity",
            "Return Outcome",
            "Request Status",
            "Due Date",
            "Returned On Time",
            "Released at",
            "Released by",
            "Returned at",
            "Returned by",
        ]

        from systems.inventory.schemas.borrow_request_schemas import BorrowRequestBatchRead
        from systems.inventory.services.borrow_request_service import BorrowService

        borrow_service = BorrowService()
        serialized_batches_by_request_uuid: dict[Any, dict[str, BorrowRequestBatchRead]] = {}
        for request in requests_ordered:
            request_batch_rows = batch_rows_by_request_id.get(request.id, [])
            if not request_batch_rows:
                continue
            serialized_batches = borrow_service.serialize_assigned_batches(
                session,
                request,
                assigned_batches=[assignment for assignment, _, _ in request_batch_rows],
            )
            serialized_batches_by_request_uuid[request.id] = {
                assignment.borrow_batch_id: assignment for assignment in serialized_batches
            }

        detail_rows: list[list[Any]] = []
        borrower_rows_map: dict[str, list[list[Any]]] = defaultdict(list)
        trackable_rows: list[list[Any]] = []
        untrackable_rows: list[list[Any]] = []

        for request in requests_ordered:
            borrower_label = self._build_borrower_label(
                user_map, request.borrower_uuid, request.customer_name,
            )
            due_date = self._format_optional_timestamp(request.return_at)
            returned_on_time_label = (
                "Yes" if request.returned_on_time is True else "No" if request.returned_on_time is False else "N/A"
            )
            request_items = request_items_by_request_id.get(request.id, [])
            request_unit_rows = unit_rows_by_request_id.get(request.id, [])
            request_batch_rows = batch_rows_by_request_id.get(request.id, [])

            for request_item in request_items:
                inventory_item = item_by_id.get(request_item.item_uuid)
                matching_units = [
                    row for row in request_unit_rows
                    if row[1] is not None and row[1].inventory_uuid == request_item.item_uuid
                ]
                matching_batches = [
                    row for row in request_batch_rows
                    if row[1] is not None and row[1].inventory_uuid == request_item.item_uuid
                ]

                if matching_units:
                    for assignment, unit, unit_item in matching_units:
                        resolved_item = unit_item or inventory_item
                        row = [
                            request.request_date.strftime("%m/%d/%Y") if request.request_date else "",
                            request.request_id,
                            borrower_label,
                            resolved_item.name if resolved_item else "Deleted Inventory Item",
                            unit.serial_number if unit and unit.serial_number else "",
                            "",
                            "1",
                            "1" if assignment.returned_at else "0",
                            "0" if assignment.returned_at else "1",
                            "Fully Returned" if assignment.returned_at else "Not Returned",
                            request.status,
                            due_date,
                            returned_on_time_label,
                            self._format_optional_timestamp(assignment.released_at or request.released_at),
                            self._format_actor_from_map(user_map, assignment.released_by or request.released_by),
                            self._format_optional_timestamp(assignment.returned_at or request.returned_at),
                            self._format_actor_from_map(user_map, assignment.returned_by or request.returned_by),
                        ]
                        self._append_inventory_row(detail_rows, row)
                        borrower_rows_map[borrower_label].append(row)
                        trackable_rows.append(row)

                elif matching_batches:
                    for assignment, batch, batch_item in matching_batches:
                        resolved_item = batch_item or inventory_item
                        serialized_batch = serialized_batches_by_request_uuid.get(request.id, {}).get(
                            assignment.borrow_batch_id
                        )
                        row = [
                            request.request_date.strftime("%m/%d/%Y") if request.request_date else "",
                            request.request_id,
                            borrower_label,
                            resolved_item.name if resolved_item else "Deleted Inventory Item",
                            batch.batch_id if batch else "",
                            resolved_item.unit_of_measure if resolved_item else "",
                            format_quantity(assignment.qty_assigned),
                            format_quantity(serialized_batch.qty_returned if serialized_batch else 0),
                            format_quantity(serialized_batch.qty_not_returned if serialized_batch else assignment.qty_assigned),
                            (
                                "Fully Returned"
                                if (serialized_batch.qty_not_returned if serialized_batch else assignment.qty_assigned) == 0
                                else "Partially Returned"
                                if (serialized_batch.qty_returned if serialized_batch else 0) > 0
                                else "Not Returned"
                            ),
                            request.status,
                            due_date,
                            returned_on_time_label,
                            self._format_optional_timestamp(assignment.released_at or request.released_at),
                            self._format_actor_from_map(user_map, request.released_by),
                            self._format_optional_timestamp(assignment.returned_at or request.returned_at),
                            self._format_actor_from_map(user_map, request.returned_by),
                        ]
                        self._append_inventory_row(detail_rows, row)
                        borrower_rows_map[borrower_label].append(row)
                        untrackable_rows.append(row)

                else:
                    row = [
                        request.request_date.strftime("%m/%d/%Y") if request.request_date else "",
                        request.request_id,
                        borrower_label,
                        inventory_item.name if inventory_item else "Deleted Inventory Item",
                        "",
                        inventory_item.unit_of_measure if inventory_item and not inventory_item.is_trackable else "",
                        format_quantity(request_item.qty_requested),
                        format_quantity(request_item.qty_requested if request.returned_at else 0),
                        format_quantity(0 if request.returned_at else request_item.qty_requested),
                        "Fully Returned" if request.returned_at else "Not Returned",
                        request.status,
                        due_date,
                        returned_on_time_label,
                        self._format_optional_timestamp(request.released_at),
                        self._format_actor_from_map(user_map, request.released_by),
                        self._format_optional_timestamp(request.returned_at),
                        self._format_actor_from_map(user_map, request.returned_by),
                    ]
                    self._append_inventory_row(detail_rows, row)
                    borrower_rows_map[borrower_label].append(row)
                    if inventory_item and inventory_item.is_trackable:
                        trackable_rows.append(row)
                    else:
                        untrackable_rows.append(row)

        report_title = self._build_borrow_history_title(
            borrower_label=selected_borrower_label,
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
        )

        if format == "csv":
            return self._create_response(headers, detail_rows, format, report_title)

        if borrower_id:
            safe_sheets: list[tuple[str, list[str], list[list[Any]], str]] = []
            if trackable_rows:
                safe_sheets.append(("Trackable/Equipments", headers, trackable_rows, report_title))
            if untrackable_rows:
                safe_sheets.append(("Untrackable/Materials", headers, untrackable_rows, report_title))
            if not safe_sheets:
                safe_sheets.append(("No Results", headers, [], report_title))
            return self._create_multi_sheet_response(safe_sheets, report_title)

        safe_sheets: list[tuple[str, list[str], list[list[Any]], str]] = []
        for borrower_label, rows in sorted(borrower_rows_map.items()):
            safe_name = borrower_label[:31]
            safe_sheets.append((safe_name, headers, rows, report_title))
        if not safe_sheets:
            safe_sheets.append(("No Results", headers, [], report_title))
        return self._create_multi_sheet_response(safe_sheets, report_title)

    def _export_movements_v2(
        self,
        session: Session,
        format: str,
        movement_type: Optional[str],
        item_id: Optional[str],
        serial_number: Optional[str],
        timeline_mode: TimelineMode | None,
        anchor_date: date | None,
        date_from: datetime | None,
        date_to: datetime | None,
        include_deleted: bool,
        include_archived: bool,
    ) -> StreamingResponse:
        if not item_id:
            raise ValueError("item_id is required for Equipment History export")

        from systems.admin.models.user import User

        normalized_serial_number = serial_number.strip() if serial_number else None
        borrower_user = aliased(User)
        selected_item = session.exec(
            self._apply_visibility_filters(
                select(InventoryItem).where(InventoryItem.item_id == item_id),
                InventoryItem,
                include_deleted,
                include_archived,
            )
        ).first()
        selected_item_label = (
            selected_item.name if selected_item and selected_item.name else item_id
        )

        statement = (
            select(BorrowRequestUnit, BorrowRequest, InventoryUnit, InventoryItem, borrower_user)
            .join(BorrowRequest, BorrowRequestUnit.borrow_uuid == BorrowRequest.id)
            .outerjoin(InventoryUnit, BorrowRequestUnit.unit_uuid == InventoryUnit.id)
            .outerjoin(InventoryItem, InventoryUnit.inventory_uuid == InventoryItem.id)
            .outerjoin(borrower_user, BorrowRequest.borrower_uuid == borrower_user.id)
        )
        statement = self._apply_visibility_filters(statement, BorrowRequestUnit, include_deleted, include_archived)
        statement = self._apply_visibility_filters(statement, BorrowRequest, include_deleted, include_archived)
        statement = self._apply_visibility_filters(statement, InventoryUnit, include_deleted, include_archived)
        statement = self._apply_visibility_filters(statement, InventoryItem, include_deleted, include_archived)
        statement = self._apply_datetime_window(statement, BorrowRequest.request_date, date_from, date_to)
        statement = statement.where(InventoryItem.item_id == item_id)

        if normalized_serial_number:
            statement = statement.where(InventoryUnit.serial_number == normalized_serial_number)

        results = self._execute_bounded_query(
            session,
            statement.order_by(InventoryUnit.serial_number.asc(), BorrowRequest.request_date.desc()),
            self._MAX_QUERY_ROWS,
            "equipment history",
        )

        actor_ids: set[Any] = set()
        for _, request, _, _, borrower in results:
            actor_ids.update({request.borrower_uuid})
        user_map = self._build_user_map(session, actor_ids)

        headers = [
            "Serial Number",
            "Request ID Reference",
            "Request Date",
            "Who Borrowed",
            "Status on Release",
            "Released at",
            "Returned at",
            "Status on Return",
        ]

        detail_rows: list[list[Any]] = []
        serial_rows_map: dict[str, list[list[Any]]] = defaultdict(list)

        for assignment, request, unit, item, borrower in results:
            if unit is None or not unit.serial_number:
                continue

            borrower_label = self._build_borrower_label(
                user_map, request.borrower_uuid, request.customer_name,
            )

            row = [
                unit.serial_number or "",
                request.request_id,
                request.request_date.strftime("%m/%d/%Y") if request.request_date else "",
                borrower_label,
                assignment.condition_on_release or "",
                self._format_optional_timestamp(assignment.released_at or request.released_at),
                self._format_optional_timestamp(assignment.returned_at or request.returned_at),
                assignment.condition_on_return or "",
            ]
            self._append_inventory_row(detail_rows, row)
            serial_rows_map[unit.serial_number].append(row)

        report_title = self._build_equipment_history_title(
            item_label=selected_item_label,
            serial_number=normalized_serial_number,
            timeline_mode=timeline_mode,
            anchor_date=anchor_date,
        )

        if format == "csv":
            return self._create_response(headers, detail_rows, format, report_title)

        safe_sheets: list[tuple[str, list[str], list[list[Any]], str]] = []
        for serial_number_value, rows in sorted(
            serial_rows_map.items(),
            key=lambda item: self._sanitize_sheet_title(item[0]),
        ):
            safe_name = serial_number_value[:31]
            safe_sheets.append((safe_name, headers, rows, report_title))
        if not safe_sheets:
            safe_sheets.append(("No Results", headers, [], report_title))
        return self._create_multi_sheet_response(safe_sheets, report_title)

    def export_entrusted(
        self,
        session: Session,
        format: str,
        search: Optional[str] = None,
        status: Optional[str] = None,
        category: Optional[str] = None,
        classification: Optional[str] = None,
    ) -> StreamingResponse:
        from systems.inventory.services.entrusted_item_service import EntrustedItemService
        
        service = EntrustedItemService()
        assignments, _ = service.get_all_entrusted(
            session=session,
            skip=0,
            limit=self._MAX_EXPORT_ROWS,
            search=search,
            status=status,
            category=category,
            classification=classification
        )

        headers = [
            "Assignment ID",
            "Item Name",
            "Serial Number",
            "Assigned To Name",
            "Assigned To ID",
            "Assigned At",
            "Returned At",
            "Status",
            "Notes"
        ]

        data = [
            [
                ass.assignment_id,
                ass.item_name or "N/A",
                ass.serial_number or "",
                ass.assigned_to_name or "Unknown User",
                ass.assigned_to_user_id,
                ass.assigned_at,
                ass.returned_at if ass.returned_at else "N/A",
                "Returned" if ass.returned_at else "Assigned",
                ass.notes or ""
            ]
            for ass in assignments
        ]

        return self._create_response(headers, data, format, "Entrusted Items Report")

    def _create_response(self, headers: List[str], data: List[List[Any]], format: str, report_title: str) -> StreamingResponse:
        filename = self._make_export_filename(report_title)
        sanitized_headers = [self._sanitize_export_cell(header) for header in headers]
        sanitized_data = self._sanitize_export_rows(data)
        
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(sanitized_headers)
            writer.writerows(sanitized_data)
            output.seek(0)
            payload = output.getvalue().encode("utf-8")
            return StreamingResponse(
                self._stream_bytes(payload),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
            )
        
        elif format == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = self._sanitize_sheet_title(report_title)
            self._populate_sheet(ws, report_title, sanitized_headers, sanitized_data)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                self._stream_bytes(output.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
            )
        
        raise ValueError(
            f"Unsupported export format: {format}. Supported formats: csv, xlsx"
        )
